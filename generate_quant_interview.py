#!/usr/bin/env python3
"""
Generate a comprehensive quant interview script in Excel format
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_quant_interview_xlsx():
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    category_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    category_font = Font(bold=True, size=11)
    answer_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Mental Math
    ws1 = wb.create_sheet("Mental Math")
    ws1.append(["#", "Question", "Answer", "Time (sec)", "Difficulty"])
    
    mental_math_questions = [
        ("1", "Calculate: 37 × 43", "1,591", "30", "Easy"),
        ("2", "What is 15% of 680?", "102", "20", "Easy"),
        ("3", "Calculate: √(2025)", "45", "20", "Easy"),
        ("4", "What is 247 ÷ 13?", "19", "30", "Medium"),
        ("5", "Calculate: 89² - 88²", "177 (use difference of squares: (89+88)(89-88))", "40", "Medium"),
        ("6", "What is 3.7 × 2.3?", "8.51", "30", "Medium"),
        ("7", "Calculate: (125 × 64) ÷ 16", "500", "45", "Medium"),
        ("8", "What is log₂(256)?", "8", "20", "Easy"),
        ("9", "Calculate: 17³", "4,913", "60", "Hard"),
        ("10", "If you invest $1000 at 7% compound annually for 3 years, what's the final amount?", "$1,225.04 (1000 × 1.07³)", "60", "Medium"),
    ]
    
    for question in mental_math_questions:
        ws1.append(question)
    
    # Sheet 2: Probability
    ws2 = wb.create_sheet("Probability")
    ws2.append(["#", "Question", "Answer/Solution", "Difficulty"])
    
    probability_questions = [
        ("1", "You flip a fair coin 3 times. What's the probability of getting exactly 2 heads?", "3/8 or 0.375\nSolution: C(3,2) × (0.5)² × (0.5)¹ = 3 × 0.125", "Easy"),
        ("2", "You roll two dice. What's the probability their sum is 7?", "1/6 or ~16.67%\nSolution: 6 ways (1+6, 2+5, 3+4, 4+3, 5+2, 6+1) out of 36 total", "Easy"),
        ("3", "In a deck of 52 cards, you draw 2 cards without replacement. What's the probability both are aces?", "1/221 or ~0.45%\nSolution: (4/52) × (3/51) = 12/2652", "Medium"),
        ("4", "You're on a game show with 3 doors. Behind one is a car, behind the others are goats. After you pick, the host opens another door with a goat. Should you switch?", "YES - Switch! Probability increases from 1/3 to 2/3\nThis is the famous Monty Hall problem", "Medium"),
        ("5", "A test is 99% accurate (99% TP, 99% TN). 0.1% of population has disease. You test positive. What's the probability you have it?", "~9.02%\nSolution: Use Bayes' Theorem\nP(D|+) = (0.001 × 0.99)/(0.001 × 0.99 + 0.999 × 0.01)", "Hard"),
        ("6", "Expected value: You pay $5 to roll a die. You win $30 if you roll a 6, otherwise nothing. Is this game favorable?", "NO\nE[V] = (1/6)(30) + (5/6)(0) - 5 = 5 - 5 = $0 (break-even)", "Medium"),
        ("7", "You have 2 envelopes, one has 2x the other. You pick one with $100. Should you switch?", "PARADOX - No advantage\nThe problem is ill-defined. Expected values don't work here.", "Hard"),
        ("8", "100 people in a room. What's the probability at least 2 share a birthday?", "~99.99997%\nWith 100 people, it's almost certain (birthday paradox)", "Medium"),
        ("9", "You draw cards until you get an ace. What's the expected number of draws?", "10.6 cards\nSolution: E[X] = (52+1)/5 = 53/5", "Hard"),
        ("10", "Two friends agree to meet between 1-2pm, each arriving at random. Each waits 15 min. Probability they meet?", "7/16 or 43.75%\nSolution: Geometric probability - area where |x-y| ≤ 15 divided by total area", "Hard"),
    ]
    
    for question in probability_questions:
        ws2.append(question)
    
    # Sheet 3: Market Making & Trading
    ws3 = wb.create_sheet("Market Making & Trading")
    ws3.append(["#", "Question", "Answer/Solution", "Difficulty"])
    
    trading_questions = [
        ("1", "What is a market maker's primary role?", "Provide liquidity by continuously quoting bid and ask prices, profiting from the bid-ask spread while managing inventory risk", "Easy"),
        ("2", "If you're asked to make a market in a stock, what factors do you consider?", "- Volatility\n- Current inventory position\n- Order flow toxicity\n- Bid-ask spread in the market\n- Adverse selection risk\n- Overall market conditions", "Medium"),
        ("3", "You're long 1000 shares of a stock trading at $50. Volatility increases. How do you adjust your quotes?", "WIDEN the spread to account for increased risk\nMay also skew quotes to reduce inventory (lower bid/ask to encourage selling)", "Medium"),
        ("4", "Explain adverse selection in market making", "Risk that informed traders trade against you, leaving you with bad positions. Market makers lose to informed flow and profit from uninformed flow.", "Medium"),
        ("5", "You quote 99-101 in a stock. Someone hits your bid for 500 shares at 99. What's your PnL?", "Unrealized loss: bought at 99, mid is 100, so -$500 (500 × (99-100))\nYou're now long 500 shares and exposed to market risk", "Easy"),
        ("6", "How would you value a European call option intuitively?", "Consider:\n- Intrinsic value: max(S-K, 0)\n- Time value: volatility × time\n- Higher for: higher S, higher σ, longer T, lower K\n(Black-Scholes gives precise value)", "Medium"),
        ("7", "You're market making in options. Implied vol is 20% but you think realized vol will be 25%. What do you do?", "BUY options (buy gamma)\nOptions are underpriced relative to expected realized volatility\nProfit from volatility arbitrage", "Medium"),
        ("8", "Explain delta hedging", "Maintaining a neutral position to changes in underlying price by offsetting delta. If long call (delta +0.5), short 0.5 shares to be delta neutral.", "Medium"),
        ("9", "Stock is at $100. You sell a $105 call for $3. At expiration, stock is at $110. What's your PnL?", "Lost $2 per share\nSolution: Received $3 premium, but option exercised at $105 when stock is $110\nPnL = 3 - (110-105) = -$2", "Medium"),
        ("10", "What's the put-call parity relationship?", "C - P = S - K×e^(-rT)\nFor European options: Call - Put = Stock - PV(Strike)\nArbitrage opportunity if violated", "Medium"),
    ]
    
    for question in trading_questions:
        ws3.append(question)
    
    # Sheet 4: Strategy & Brainteasers
    ws4 = wb.create_sheet("Strategy & Brainteasers")
    ws4.append(["#", "Question", "Answer/Solution", "Difficulty"])
    
    strategy_questions = [
        ("1", "You have 8 balls, one is heavier. You have a balance scale. Minimum weighings to find the heavy ball?", "2 weighings\nDivide into 3-3-2. Weigh first two groups. Heavy group → weigh 2 balls from it. Can identify in 2 weighings.", "Medium"),
        ("2", "100 floor building, 2 eggs. Find highest safe floor with minimum drops in worst case?", "14 drops (worst case)\nStart at floor 14, then 27, 39, 50, 60, 69, 77, 84, 90, 95, 99, 100\nUse formula: n(n+1)/2 ≥ 100", "Hard"),
        ("3", "You and a friend take turns placing quarters on a round table. First unable to place loses. Who wins with optimal play?", "First player wins\nStrategy: Place first quarter in exact center, then mirror opponent's moves", "Medium"),
        ("4", "25 horses, 5 race tracks. Find top 3 fastest with minimum races?", "7 races\nRun 5 heats (5 races). Race winners (1 race). Eliminate and race necessary horses (1 race).", "Hard"),
        ("5", "You have $100. Coin flip: heads you win $50, tails you lose $40. How much should you bet each round to maximize long-term growth?", "Kelly Criterion: f* = (p×b - q)/b\nHere: p=0.5, b=1.25, q=0.5\nf* = (0.5×1.25 - 0.5)/1.25 = 12.5% per round\nBet $12.50 each time", "Hard"),
        ("6", "Design a trading strategy for mean-reverting vs trending markets", "Mean-reverting: Sell high, buy low. Use Bollinger Bands, RSI. Fade extremes.\nTrending: Buy high, sell higher. Use moving averages, breakouts. Follow momentum.\nKey: Identify regime first!", "Medium"),
        ("7", "You observe a coin flip 10 times: 7 heads, 3 tails. Estimate probability of heads on next flip?", "Bayesian approach: ~0.571 (posterior mean with uniform prior)\nOr use (successes + 1)/(trials + 2) = 8/12 = 0.667\nClassical: 0.7 (but small sample)", "Medium"),
        ("8", "Explain how you'd detect a rigged roulette wheel", "Statistical approach:\n1. Collect large sample (>1000 spins)\n2. Chi-square test for uniform distribution\n3. Look for sector bias (quadrant analysis)\n4. Test for sequential correlation\n5. Compare to expected variance", "Medium"),
        ("9", "Why are manhole covers round?", "Multiple reasons:\n- Can't fall through the hole (no orientation where diameter < hole)\n- Easy to roll\n- No alignment needed\n- Structural strength uniform", "Easy"),
        ("10", "How would you hedge a book of tech call options during earnings season?", "Multiple approaches:\n- Delta hedge with underlying stock\n- Vega hedge with other options (buy puts for volatility hedge)\n- Calendar spread to manage time decay\n- Sector ETF hedge (QQQ, XLK)\n- Consider correlation with other tech stocks", "Hard"),
    ]
    
    for question in strategy_questions:
        ws4.append(question)
    
    # Sheet 5: Statistics & Math
    ws5 = wb.create_sheet("Statistics & Math")
    ws5.append(["#", "Question", "Answer/Solution", "Difficulty"])
    
    stats_questions = [
        ("1", "What's the difference between correlation and covariance?", "Covariance: Cov(X,Y) = E[(X-μₓ)(Y-μᵧ)] (units: product of variables)\nCorrelation: ρ = Cov(X,Y)/(σₓσᵧ) (unitless, range: [-1, 1])\nCorrelation is normalized covariance", "Easy"),
        ("2", "Explain the Central Limit Theorem and why it matters in finance", "Sample means approach normal distribution as n→∞, regardless of underlying distribution.\nImportance: Justifies normal models for returns, enables statistical inference, underpins risk models", "Medium"),
        ("3", "What is the difference between Type I and Type II errors?", "Type I (α): False Positive - Reject true null hypothesis\nType II (β): False Negative - Fail to reject false null hypothesis\nPower = 1-β", "Easy"),
        ("4", "Calculate the variance of a portfolio: 60% Stock A (σ=20%), 40% Stock B (σ=30%), correlation=0.5", "σₚ² = wₐ²σₐ² + wᵦ²σᵦ² + 2wₐwᵦρσₐσᵦ\n= 0.36×400 + 0.16×900 + 2×0.6×0.4×0.5×20×30\n= 144 + 144 + 144 = 432\nσₚ = 20.78%", "Medium"),
        ("5", "What is the Sharpe Ratio and how do you interpret it?", "Sharpe = (E[R] - Rf) / σ\nMeasures risk-adjusted return (excess return per unit of risk)\n>1 is good, >2 is very good, >3 is excellent", "Easy"),
        ("6", "Derive the expected value of the maximum of two uniform [0,1] random variables", "E[max(X,Y)] = ∫∫max(x,y)dxdy = ∫₀¹∫₀¹max(x,y)dxdy = 2/3\nAlternatively: E[max] = 2/(n+1) for n=2 gives 2/3", "Hard"),
        ("7", "What is Jensen's inequality? Give a finance example.", "For convex f: E[f(X)] ≥ f(E[X])\nFor concave f: E[f(X)] ≤ f(E[X])\nExample: Log utility is concave, so E[log(W)] ≤ log(E[W])\nImplies investors prefer certainty to risky gambles with same expected value", "Medium"),
        ("8", "If X ~ N(0,1), what is E[X³]?", "0\nOdd moments of symmetric distributions around mean are 0\nNormal distribution is symmetric around mean", "Medium"),
        ("9", "Explain heteroskedasticity and why it matters", "Non-constant variance across observations\nProblems: OLS estimators inefficient, standard errors biased, hypothesis tests invalid\nSolutions: GLS, robust standard errors, GARCH models in finance", "Medium"),
        ("10", "You have two dice. What's the expected value of the maximum?", "E[max(X,Y)] = 4.472 (≈161/36)\nMethod: P(max≤k) = (k/6)², so E[max] = Σk×P(max=k)\nOr use: E[max] = Σk=1⁶ k×[(k²-(k-1)²)/36]", "Hard"),
    ]
    
    for question in stats_questions:
        ws5.append(question)
    
    # Sheet 6: Coding & Algorithms
    ws6 = wb.create_sheet("Coding & Algorithms")
    ws6.append(["#", "Question", "Approach/Pseudocode", "Difficulty"])
    
    coding_questions = [
        ("1", "Implement a function to calculate a moving average of a time series", "Use deque or circular buffer:\n- Initialize window of size k\n- Add new element, remove oldest\n- Maintain running sum\nTime: O(1) per update, Space: O(k)", "Easy"),
        ("2", "Find the maximum profit from a stock price array (buy once, sell once)", "One pass algorithm:\ntrack min_price, max_profit\nfor each price:\n  profit = price - min_price\n  max_profit = max(max_profit, profit)\n  min_price = min(min_price, price)\nTime: O(n), Space: O(1)", "Medium"),
        ("3", "Implement a LRU cache with O(1) get and put", "Use HashMap + Doubly Linked List:\n- HashMap: key → node\n- DLL: maintain recency order\nget(): Move to front\nput(): Add to front, evict tail if full\nTime: O(1) both operations", "Medium"),
        ("4", "Calculate Value at Risk (VaR) for a portfolio", "Historical simulation approach:\n1. Calculate historical returns\n2. Sort returns ascending\n3. VaR = percentile(returns, α)\n   e.g., 95% VaR = 5th percentile\nAlternatively: Parametric (assume normal)", "Medium"),
        ("5", "Implement Black-Scholes option pricing", "C = S×N(d₁) - K×e^(-rT)×N(d₂)\nd₁ = [ln(S/K) + (r + σ²/2)T] / (σ√T)\nd₂ = d₁ - σ√T\n\nNeed: cumulative normal distribution function\nLibraries: scipy.stats.norm.cdf", "Hard"),
        ("6", "Detect arbitrage in currency exchange rates (Bellman-Ford)", "Model as graph:\n- Vertices: currencies\n- Edges: exchange rates\n- Use log rates: log(rate)\n- Find negative cycle\nNegative cycle = arbitrage opportunity\nTime: O(VE)", "Hard"),
        ("7", "Design a system to handle high-frequency tick data", "Architecture:\n- Message queue (Kafka/RabbitMQ)\n- In-memory store (Redis)\n- Time-series DB (InfluxDB, TimescaleDB)\n- Stream processing (Flink, Kafka Streams)\n- Low-latency considerations\n- Lock-free data structures", "Hard"),
        ("8", "Implement a Monte Carlo option pricer", "1. Generate N random paths\n2. For each path:\n   - Simulate stock: S(t+1) = S(t)×exp((μ-σ²/2)dt + σ√dt×Z)\n   - Calculate payoff at maturity\n3. Average discounted payoffs\n4. Standard error: σ/√N", "Medium"),
        ("9", "Find the median of a stream of numbers efficiently", "Use two heaps:\n- Max heap (lower half)\n- Min heap (upper half)\nMedian = top of max heap (or avg of tops)\ninsert(): O(log n), getMedian(): O(1)", "Medium"),
        ("10", "Implement a simple backtesting framework structure", "Components:\n1. Data handler (load/stream data)\n2. Strategy (signal generation)\n3. Portfolio (position management)\n4. Execution (order handling)\n5. Performance metrics\n\nLoop: For each bar, generate signals, execute, track PnL", "Medium"),
    ]
    
    for question in coding_questions:
        ws6.append(question)
    
    # Sheet 7: Interview Tips
    ws7 = wb.create_sheet("Interview Tips & Structure")
    ws7.append(["Section", "Details"])
    
    tips = [
        ("Interview Format", "Typical quant onsite: 4-6 rounds\n- Mental math/probability (30-45 min)\n- Market making/trading (30-45 min)\n- Coding (45-60 min)\n- Behavioral (30 min)\n- Brain teasers (30 min)"),
        ("Mental Math Tips", "- Break down complex calculations\n- Use estimation when appropriate\n- Know squares up to 20², cubes up to 12³\n- Master percentage calculations\n- Practice under time pressure"),
        ("Probability Tips", "- Always define the sample space\n- Use tree diagrams for complex problems\n- Know common distributions\n- Bayes' theorem is crucial\n- Think about edge cases"),
        ("Market Making Tips", "- Understand bid-ask spread dynamics\n- Discuss inventory risk management\n- Know Greeks (delta, gamma, vega, theta)\n- Explain adverse selection\n- Consider market microstructure"),
        ("Strategy Tips", "- Think out loud\n- Ask clarifying questions\n- Start simple, then optimize\n- Discuss trade-offs\n- Consider real-world constraints"),
        ("Coding Tips", "- Clarify requirements first\n- Discuss time/space complexity\n- Test edge cases\n- Write clean, readable code\n- Explain your approach before coding"),
        ("Key Formulas", "• Expected Value: E[X] = Σ x×P(x)\n• Variance: Var(X) = E[X²] - (E[X])²\n• Covariance: Cov(X,Y) = E[XY] - E[X]E[Y]\n• Sharpe: (μ-rf)/σ\n• Black-Scholes: C = S×N(d₁) - K×e^(-rT)×N(d₂)\n• Kelly: f* = (p×b - q)/b"),
        ("Common Distributions", "• Uniform [a,b]: E[X]=(a+b)/2, Var=((b-a)²/12)\n• Normal N(μ,σ²): E[X]=μ, Var=σ²\n• Exponential λ: E[X]=1/λ, Var=1/λ²\n• Poisson λ: E[X]=λ, Var=λ\n• Binomial n,p: E[X]=np, Var=np(1-p)"),
        ("Things to Review", "☐ Options pricing (Black-Scholes)\n☐ Greek letters and hedging\n☐ Common probability distributions\n☐ Bayes' theorem applications\n☐ Market microstructure\n☐ Risk metrics (VaR, Sharpe, Sortino)\n☐ Time series analysis\n☐ Stochastic calculus basics\n☐ Monte Carlo simulation\n☐ Statistical arbitrage"),
        ("Red Flags to Avoid", "✗ Jumping to solution without understanding\n✗ Not checking edge cases\n✗ Overcomplicating simple problems\n✗ Not communicating your thought process\n✗ Giving up too quickly\n✗ Being overconfident without verification\n✗ Not asking questions when unclear"),
    ]
    
    for tip in tips:
        ws7.append(tip)
    
    # Apply formatting to all sheets
    for ws in wb.worksheets:
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        
        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set specific widths for readability
        if ws.title == "Mental Math":
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 50
        elif ws.title in ["Probability", "Market Making & Trading", "Strategy & Brainteasers", "Statistics & Math", "Coding & Algorithms"]:
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 70
        elif ws.title == "Interview Tips & Structure":
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 100
        
        # Freeze first row
        ws.freeze_panes = ws['A2']
    
    # Save workbook
    wb.save('Quant_Interview_Script.xlsx')
    print("✅ Excel file 'Quant_Interview_Script.xlsx' created successfully!")
    print("\nSummary:")
    print("- Sheet 1: Mental Math (10 questions)")
    print("- Sheet 2: Probability (10 questions)")
    print("- Sheet 3: Market Making & Trading (10 questions)")
    print("- Sheet 4: Strategy & Brainteasers (10 questions)")
    print("- Sheet 5: Statistics & Math (10 questions)")
    print("- Sheet 6: Coding & Algorithms (10 questions)")
    print("- Sheet 7: Interview Tips & Structure")
    print("\nTotal: 60 interview questions + comprehensive tips")

if __name__ == "__main__":
    create_quant_interview_xlsx()
